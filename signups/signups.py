import pandas as pd
from openpyxl import load_workbook
import datetime
from utils.sheet_formatting import *
from utils.from_sftp import *

def signups():
    '''generate an excel file using the signups_template'''
    # read necessary tables
    deas = pd.read_csv('data/az_prescriber_deas.csv', index_col=None)
    awarxe = pd.read_excel('data/awarxe.xls', skiprows=1, index_col=None)
    exclude_dvm = pd.read_csv('data/exclude_dvm.csv', index_col=None)
    city_county_lkup = pd.read_csv('data/lookup_city_county.csv', index_col=None)
    typo_city_lkup = pd.read_csv('data/lookup_mispelled_city.csv', index_col=None)

    # get the previous month
    today = datetime.date.today()
    first = today.replace(day=1)
    last_month_day = first - datetime.timedelta(days=1)
    last_month_name = last_month_day.strftime('%B')
    last_month_num = last_month_day.strftime('%m')
    last_month_year = last_month_day.strftime('%Y')
    last_month_str = f'{last_month_name} {last_month_year}'

    # get the numbers of the current month and year
    cur_month = today.strftime('%m')
    cur_year = today.strftime('%Y')

    # format the date for sftp file name
    last_my = f'{int(cur_year) - 1}{cur_month}'
    cur_my = f'{last_month_year}{last_month_num}'
    print(f'getting {last_my} - {cur_my} data')

    dispensations = from_sftp(f'/Monthly/REPORTAE_47/Dispensations_12_Months/AZ_Dispensations_{last_my}_{cur_my}.csv')

    # output file path
    name_str = f'data/signups{cur_month}01{cur_year}.xlsx'

    dispensations['state'] = dispensations['state'].str.upper()

    # create az_pro_info from dispensations and DEA list
    az_dispensations = dispensations[(dispensations['state'] == 'AZ') | (dispensations['state'] == 'ARIZONA')]
    az_dispensations = az_dispensations.drop(az_dispensations.dropna(subset=['dea_suffix']).index)
    pro_info_sum = az_dispensations[['dea_number', 'rx_count']].groupby('dea_number', as_index=False).sum()
    az_pro_info = pd.merge(pro_info_sum, deas[['DEA Number', 'Name', 'Additional Company Info', 'Address 1', 'Address 2', 'City', 'State', 'Zip Code']], 
        left_on='dea_number', right_on='DEA Number', how='inner').drop(columns=['DEA Number'])

    # fix typos in city and add county
    az_pro_info.rename(columns={'City':'pi_city'}, inplace=True)
    az_pro_info['pi_city'] = az_pro_info['pi_city'].str.upper()
    az_pro_info = pd.merge(az_pro_info, typo_city_lkup[['CityError', 'City']], 
        left_on='pi_city', right_on='CityError', how='left')
    az_pro_info['fixed_city'] = az_pro_info['City'].fillna(az_pro_info['pi_city'])
    az_pro_info = az_pro_info.drop(columns=['pi_city','CityError','City'])
    az_pro_info.rename(columns={'fixed_city':'City'}, inplace=True)
    az_pro_info = pd.merge(az_pro_info, city_county_lkup[['City', 'County']], how='left')

    # make a list of prescribers with no county
    no_county = az_pro_info[az_pro_info['County'].isna()]
    no_county.to_clipboard(index=False)
    print('prescribers with no county have been copied to clipboard')

    # keep only prescribers with a county
    az_pro_info = az_pro_info[az_pro_info['County'].notna()]

    # update exclude_dvm
    vets = az_pro_info[az_pro_info['Name'].str.contains('DVM|VMD')]
    vets = vets[['dea_number']]
    vets.rename(columns={'dea_number':'DEA'}, inplace=True)
    new_vets = vets.merge(exclude_dvm, how='left', indicator=True)
    new_vets = new_vets[new_vets['_merge'] == 'left_only'][['DEA']]
    new_exclude_dvm = pd.concat([exclude_dvm, new_vets])
    new_exclude_dvm.to_csv('data/exclude_dvm.csv', index=False)

    # drop vets from list
    az_pro_info = az_pro_info.merge(new_exclude_dvm, left_on='dea_number', right_on='DEA', how='left', indicator=True)
    az_pro_info = az_pro_info[az_pro_info['_merge'] == 'left_only'].drop(columns=['DEA', '_merge'])

    # drop hospital style names
    az_pro_info = az_pro_info[~az_pro_info['Name'].str.contains('CENTER|HOSPITAL|SCOTTSDALE|MEDICAL|REGIONAL')]

    # check registration in awarxe
    az_pro_info = az_pro_info.assign(awarxe=az_pro_info['dea_number'].isin(awarxe['DEA Number']))
    az_pro_info['awarxe'] = az_pro_info['awarxe'].map({True:'YES' ,False:'NO'})
    az_pro_info.drop_duplicates(subset="dea_number", keep='first', inplace=True)

    # rearrange and sort the table
    az_pro_info = az_pro_info[['awarxe', 'County', 'Name', 'Additional Company Info', 'Address 1', 'Address 2', 'City', 'State', 'Zip Code']]
    az_pro_info.sort_values(['County'], inplace=True)
    az_pro_info.rename(columns={'awarxe':'AWARxE Account?', 'Name':'Prescriber Name'}, inplace=True)

    # make the totals table
    yeses = az_pro_info[['County', 'AWARxE Account?']]
    yeses = yeses[yeses['AWARxE Account?'] == 'YES']
    yeses = yeses.groupby('County', as_index=False).count()
    county_count = az_pro_info[['County']].value_counts().reset_index(name='county_total')
    totals = pd.merge(yeses, county_count, on='County', how='inner')

    # get delegate users in casa grande
    awarxe_delegates = awarxe[awarxe['Role Title'].str.contains('Delegate') | awarxe['Role Title'].str.contains('Technician')]
    awarxe_delegates = awarxe_delegates[awarxe_delegates['City'].str.upper() == 'CASA GRANDE']
    awarxe_delegates_presc = awarxe_delegates[awarxe_delegates['Role Title'].str.contains('Presc')]
    awarxe_delegates_pharm = awarxe_delegates[awarxe_delegates['Role Title'].str.contains('Pharm')]
    print(f'number of CG prescriber delegates: {len(awarxe_delegates_presc)}')
    print(f'number of CG pharmacist delegates: {len(awarxe_delegates_pharm)}')

    # write to excel
    template_wb = load_workbook('data/signups_template.xlsx')
    totals_ws = template_wb['Totals']
    totals_ws['A24'] = last_month_str
    totals_ws['C21'] = len(awarxe_delegates_presc)
    totals_ws['D21'] = len(awarxe_delegates_pharm)

    totals_np = totals.to_numpy()
    for i in range(3,18):
        j = i - 3
        totals_ws[f'B{i}'] = totals_np[j][1]
        totals_ws[f'C{i}'] = totals_np[j][2]

    template_wb.save(name_str)

    book = load_workbook(name_str)
    writer = pd.ExcelWriter(name_str, engine='openpyxl')
    writer.book = book
    az_pro_info.style.applymap(highlight_cells).to_excel(writer, index=False, sheet_name='County Lists', engine='openpyxl')
    set_col_widths_openpyxl(writer, az_pro_info, 'County Lists')
    writer.save()

def main():
    print('starting signups')
    signups()
    print('signups saved')

if __name__ == "__main__":
    main()